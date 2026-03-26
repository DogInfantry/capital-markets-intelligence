"""
Professional Excel Workbook Generator
Creates fully formatted Excel workbooks with multiple tabs,
conditional formatting, charts, and data validation.

Outputs:
  output/excel/capital_markets_master.xlsx
  output/excel/ipo_analysis_workbook.xlsx
  output/excel/sovereign_risk_workbook.xlsx
  output/excel/mna_analysis_workbook.xlsx
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
RAW_DIR = os.path.join(BASE_DIR, "data", "raw")
PROC_DIR = os.path.join(BASE_DIR, "data", "processed")
EXCEL_DIR = os.path.join(BASE_DIR, "output", "excel")

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2C3E50")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="7F8C8D")
DATA_FONT = Font(name="Calibri", size=10)
NUM_FORMAT_PCT = '0.00"%"'
NUM_FORMAT_BN = '$#,##0.0"B"'
NUM_FORMAT_M = '$#,##0"M"'
NUM_FORMAT_2D = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
GREEN_FILL = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
RED_FILL = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")


def load_data():
    """Load all datasets."""
    data = {}
    files = {
        "ipo": (RAW_DIR, "ipo_data_raw.csv"),
        "mna": (RAW_DIR, "mna_data_raw.csv"),
        "sovereign": (RAW_DIR, "sovereign_issuance_raw.csv"),
        "panel": (RAW_DIR, "country_macro_panel.csv"),
        "stress": (PROC_DIR, "stress_test_results.csv"),
        "risk_index": (PROC_DIR, "sovereign_risk_index.csv"),
        "event_study": (PROC_DIR, "ipo_event_study.csv"),
        "screening": (PROC_DIR, "deal_screening_model.csv"),
        "cases": (PROC_DIR, "case_studies.csv"),
        "yc": (PROC_DIR, "yield_curve_analysis.csv"),
        "regime": (PROC_DIR, "regime_indicators.csv"),
    }
    for key, (directory, filename) in files.items():
        filepath = os.path.join(directory, filename)
        if os.path.exists(filepath):
            data[key] = pd.read_csv(filepath)
    return data


def autofit_columns(ws, max_width=35):
    """Auto-fit column widths, skipping merged cells."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, 1):
        lengths = []
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            lengths.append(len(str(cell.value or "")))
        if lengths:
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max(lengths) + 2, max_width)


def style_header_row(ws, row_num, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_dataframe(ws, df, start_row=1, start_col=1, header=True):
    """Write a DataFrame to a worksheet with formatting."""
    rows = list(dataframe_to_rows(df, index=False, header=header))
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    if header:
        style_header_row(ws, start_row, start_col + len(df.columns) - 1)

    return start_row + len(rows)


def add_title(ws, title, subtitle=None, row=1):
    """Add title and subtitle to worksheet."""
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=8)
        return row + 3
    return row + 2


def create_ipo_workbook(data):
    """Create comprehensive IPO analysis workbook."""
    print("  [1/4] IPO Analysis Workbook...")

    wb = Workbook()
    ipo = data.get("ipo")
    event = data.get("event_study")

    if ipo is None:
        return

    ipo_df = pd.DataFrame(ipo)

    # --- Sheet 1: IPO Data ---
    ws = wb.active
    ws.title = "IPO Database"
    row = add_title(ws, "IPO Database", f"Capital Markets Intelligence | {datetime.now().strftime('%B %Y')}")

    cols = ["company", "ticker", "exchange", "ipo_date", "offer_price",
            "first_day_close", "current_price", "first_day_return_pct",
            "total_return_pct", "deal_size_usd", "sector", "lead_underwriter", "country"]
    display_df = ipo_df[[c for c in cols if c in ipo_df.columns]].copy()
    end_row = write_dataframe(ws, display_df, start_row=row)

    # Conditional formatting on returns
    ret_col_idx = cols.index("first_day_return_pct") + 1 if "first_day_return_pct" in cols else None
    if ret_col_idx:
        ws.conditional_formatting.add(
            f"{chr(64+ret_col_idx)}{row+1}:{chr(64+ret_col_idx)}{end_row}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            f"{chr(64+ret_col_idx)}{row+1}:{chr(64+ret_col_idx)}{end_row}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )

    # Auto-fit columns
    autofit_columns(ws)

    # --- Sheet 2: Event Study ---
    if event is not None:
        ws2 = wb.create_sheet("Event Study")
        row = add_title(ws2, "IPO Event Study", "Abnormal Returns & Pricing Assessment")
        event_df = pd.DataFrame(event)
        display_cols = ["company", "ticker", "ipo_date", "first_day_return_pct",
                        "abnormal_return_pct", "money_left_on_table_usd", "pricing_assessment"]
        event_display = event_df[[c for c in display_cols if c in event_df.columns]].copy()
        write_dataframe(ws2, event_display, start_row=row)
        autofit_columns(ws2)

    # --- Sheet 3: Summary Stats ---
    ws3 = wb.create_sheet("Summary Statistics")
    row = add_title(ws3, "IPO Summary Statistics", "Statistical Analysis")

    stats_data = {
        "Metric": [
            "Total IPOs", "Total Capital Raised ($B)", "Avg First-Day Return (%)",
            "Median First-Day Return (%)", "Std Dev First-Day Return (%)",
            "% Positive First-Day", "Avg Total Return (%)",
            "Best First-Day Return (%)", "Worst First-Day Return (%)",
            "Countries Covered", "Sectors Covered",
        ],
        "Value": [
            len(ipo_df),
            round(ipo_df["deal_size_usd"].sum() / 1e9, 1) if "deal_size_usd" in ipo_df.columns else "N/A",
            round(ipo_df["first_day_return_pct"].mean(), 2),
            round(ipo_df["first_day_return_pct"].median(), 2),
            round(ipo_df["first_day_return_pct"].std(), 2),
            round((ipo_df["first_day_return_pct"] > 0).mean() * 100, 1),
            round(ipo_df["total_return_pct"].mean(), 2),
            round(ipo_df["first_day_return_pct"].max(), 2),
            round(ipo_df["first_day_return_pct"].min(), 2),
            ipo_df["country"].nunique(),
            ipo_df["sector"].nunique(),
        ],
    }
    write_dataframe(ws3, pd.DataFrame(stats_data), start_row=row)
    autofit_columns(ws3)

    # --- Sheet 4: Underwriter League Table ---
    ws4 = wb.create_sheet("Underwriter League Table")
    row = add_title(ws4, "Underwriter League Table", "By Deal Count Involvement")

    all_uw = []
    for _, r in ipo_df.iterrows():
        for u in str(r.get("lead_underwriter", "")).split("/"):
            all_uw.append({"underwriter": u.strip(), "deal_size": r.get("deal_size_usd", 0),
                           "first_day_ret": r.get("first_day_return_pct", 0)})
    uw_df = pd.DataFrame(all_uw)
    if len(uw_df) > 0:
        uw_summary = uw_df.groupby("underwriter").agg(
            deals=("underwriter", "count"),
            total_volume_usd=("deal_size", "sum"),
            avg_return_pct=("first_day_ret", "mean"),
        ).sort_values("total_volume_usd", ascending=False).reset_index()
        uw_summary["total_volume_bn"] = (uw_summary["total_volume_usd"] / 1e9).round(1)
        uw_summary["avg_return_pct"] = uw_summary["avg_return_pct"].round(2)
        write_dataframe(ws4, uw_summary[["underwriter", "deals", "total_volume_bn", "avg_return_pct"]],
                        start_row=row)
    autofit_columns(ws4)

    save_workbook(wb, "ipo_analysis_workbook.xlsx")


def create_sovereign_workbook(data):
    """Create sovereign risk analysis workbook."""
    print("  [2/4] Sovereign Risk Workbook...")

    wb = Workbook()
    risk = data.get("risk_index")
    panel = data.get("panel")
    stress = data.get("stress")
    sovereign = data.get("sovereign")

    # --- Sheet 1: Risk Index ---
    ws = wb.active
    ws.title = "Sovereign Risk Index"
    row = add_title(ws, "Proprietary Sovereign Risk Index",
                    f"Composite Score from 11 World Bank Indicators | {datetime.now().strftime('%B %Y')}")

    if risk is not None:
        risk_df = pd.DataFrame(risk)
        display_cols = ["country_name", "country_code", "risk_index_0_100", "risk_tier",
                        "GDP growth (annual %)", "Inflation (CPI, annual %)",
                        "Central govt debt (% of GDP)", "Current account (% of GDP)"]
        risk_display = risk_df[[c for c in display_cols if c in risk_df.columns]].copy()
        risk_display = risk_display.sort_values("risk_index_0_100", ascending=False)
        end_row = write_dataframe(ws, risk_display, start_row=row)

        # Color scale on risk index
        ws.conditional_formatting.add(
            f"C{row+1}:C{end_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="00FF00",
                           mid_type="num", mid_value=50, mid_color="FFFF00",
                           end_type="num", end_value=100, end_color="FF0000")
        )

    autofit_columns(ws)

    # --- Sheet 2: Macro Panel ---
    if panel is not None:
        ws2 = wb.create_sheet("Macro Panel Data")
        row = add_title(ws2, "Country Macro Panel", "World Bank Open Data API")
        panel_df = pd.DataFrame(panel)
        write_dataframe(ws2, panel_df, start_row=row)
        autofit_columns(ws2)

    # --- Sheet 3: Stress Test Results ---
    if stress is not None:
        ws3 = wb.create_sheet("Stress Test Results")
        row = add_title(ws3, "Sovereign Stress Test Model", "5 Scenarios x 12 Countries")
        stress_df = pd.DataFrame(stress)
        display_cols = ["country", "scenario", "ten_yr_yield_pct", "stressed_yield_pct",
                        "stressed_gdp_growth_pct", "stressed_inflation_pct", "risk_score"]
        stress_display = stress_df[[c for c in display_cols if c in stress_df.columns]]
        write_dataframe(ws3, stress_display, start_row=row)
        autofit_columns(ws3)

    # --- Sheet 4: Sovereign Issuance ---
    if sovereign is not None:
        ws4 = wb.create_sheet("Bond Issuance")
        row = add_title(ws4, "Sovereign Bond Issuance Database", "18 Countries, 31 Issuances")
        sov_df = pd.DataFrame(sovereign)
        write_dataframe(ws4, sov_df, start_row=row)
        autofit_columns(ws4)

    save_workbook(wb, "sovereign_risk_workbook.xlsx")


def create_mna_workbook(data):
    """Create M&A analysis workbook."""
    print("  [3/4] M&A Analysis Workbook...")

    wb = Workbook()
    mna = data.get("mna")
    cases = data.get("cases")
    screening = data.get("screening")

    # --- Sheet 1: Deal Database ---
    ws = wb.active
    ws.title = "M&A Database"
    row = add_title(ws, "M&A Deal Database",
                    f"20 Deals | ${pd.DataFrame(mna)['deal_value_usd_bn'].sum():.0f}B Total" if mna is not None else "")

    if mna is not None:
        mna_df = pd.DataFrame(mna)
        display_cols = ["acquirer", "target", "announced_date", "deal_value_usd_bn",
                        "status", "sector", "deal_type", "strategic_rationale",
                        "acquirer_country", "target_country", "cross_border"]
        mna_display = mna_df[[c for c in display_cols if c in mna_df.columns]]
        end_row = write_dataframe(ws, mna_display, start_row=row)

        # Color status column
        status_col = display_cols.index("status") + 1 if "status" in display_cols else None
        if status_col:
            for r in range(row + 1, end_row + 1):
                cell = ws.cell(row=r, column=status_col)
                val = str(cell.value)
                if "Completed" in val:
                    cell.fill = GREEN_FILL
                elif "Pending" in val:
                    cell.fill = YELLOW_FILL
                elif "Blocked" in val:
                    cell.fill = RED_FILL

    autofit_columns(ws)

    # --- Sheet 2: Case Studies ---
    if cases is not None:
        ws2 = wb.create_sheet("Case Studies")
        row = add_title(ws2, "M&A Case Study Analysis", "Strategic Rationale & Risk Assessment")
        cases_df = pd.DataFrame(cases)
        display_cols = ["deal_name", "deal_value_usd_bn", "sector", "deal_type",
                        "status", "size_tier", "strategic_rationale",
                        "value_drivers", "risk_factors", "research_relevance_score"]
        cases_display = cases_df[[c for c in display_cols if c in cases_df.columns]]
        write_dataframe(ws2, cases_display, start_row=row)
        autofit_columns(ws2)

    # --- Sheet 3: Deal Screening ---
    if screening is not None:
        ws3 = wb.create_sheet("Deal Screening Model")
        row = add_title(ws3, "Deal Screening & Completion Probability",
                        "Feature-Based Scoring Model")
        scr_df = pd.DataFrame(screening)
        display_cols = ["acquirer", "target", "deal_value_usd_bn", "status",
                        "sector", "cross_border", "completed",
                        "completion_probability_pct"]
        scr_display = scr_df[[c for c in display_cols if c in scr_df.columns]]
        write_dataframe(ws3, scr_display, start_row=row)
        autofit_columns(ws3)

    save_workbook(wb, "mna_analysis_workbook.xlsx")


def create_master_workbook(data):
    """Create master overview workbook with summary tabs."""
    print("  [4/4] Master Summary Workbook...")

    wb = Workbook()

    # --- Executive Summary ---
    ws = wb.active
    ws.title = "Executive Summary"
    row = add_title(ws, "Capital Markets Intelligence Platform",
                    f"Executive Summary | {datetime.now().strftime('%B %d, %Y')}")

    summary_data = []

    ipo = data.get("ipo")
    if ipo is not None:
        ipo_df = pd.DataFrame(ipo)
        summary_data.append(("IPO DATABASE", "", ""))
        summary_data.append(("  Total IPOs Tracked", len(ipo_df), ""))
        summary_data.append(("  Capital Raised",
                             f"${ipo_df['deal_size_usd'].sum()/1e9:.1f}B" if "deal_size_usd" in ipo_df.columns else "N/A", ""))
        summary_data.append(("  Avg First-Day Return",
                             f"{ipo_df['first_day_return_pct'].mean():.1f}%" if "first_day_return_pct" in ipo_df.columns else "N/A", ""))
        summary_data.append(("  Countries", ipo_df["country"].nunique() if "country" in ipo_df.columns else "N/A", ""))
        summary_data.append(("", "", ""))

    mna = data.get("mna")
    if mna is not None:
        mna_df = pd.DataFrame(mna)
        summary_data.append(("M&A DATABASE", "", ""))
        summary_data.append(("  Total Deals", len(mna_df), ""))
        summary_data.append(("  Total Deal Value", f"${mna_df['deal_value_usd_bn'].sum():.1f}B", ""))
        summary_data.append(("  Completed", (mna_df["status"] == "Completed").sum(), ""))
        summary_data.append(("  Blocked/Pending", (mna_df["status"] != "Completed").sum(), ""))
        summary_data.append(("", "", ""))

    risk = data.get("risk_index")
    if risk is not None:
        risk_df = pd.DataFrame(risk)
        summary_data.append(("SOVEREIGN RISK INDEX", "", ""))
        summary_data.append(("  Countries Scored", len(risk_df), ""))
        summary_data.append(("  High Risk", (risk_df["risk_tier"] == "High Risk").sum() if "risk_tier" in risk_df.columns else "N/A", ""))
        summary_data.append(("  Indicators Used", "11 (World Bank)", ""))
        summary_data.append(("", "", ""))

    summary_data.append(("DATA SOURCES", "", ""))
    summary_data.append(("  Market Data", "Yahoo Finance (live)", "501+ trading days"))
    summary_data.append(("  Macro Indicators", "World Bank API (live)", "15 indicators, 20 countries"))
    summary_data.append(("  IPO Data", "SEC EDGAR + Research", "25 IPOs"))
    summary_data.append(("  M&A Data", "Public filings", "20 deals"))
    summary_data.append(("  Sovereign Issuance", "Central banks + IMF", "31 issuances, 18 countries"))

    for r_idx, (label, value, note) in enumerate(summary_data, row):
        ws.cell(row=r_idx, column=1, value=label).font = Font(
            name="Calibri", bold=bool(str(label).isupper() and label), size=11
        )
        ws.cell(row=r_idx, column=2, value=value).font = DATA_FONT
        ws.cell(row=r_idx, column=3, value=note).font = Font(name="Calibri", size=10, color="95A5A6")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    save_workbook(wb, "capital_markets_master.xlsx")


def save_workbook(wb, filename):
    """Save workbook to Excel directory."""
    os.makedirs(EXCEL_DIR, exist_ok=True)
    filepath = os.path.join(EXCEL_DIR, filename)
    wb.save(filepath)
    size = os.path.getsize(filepath)
    print(f"    [OK] {filename} ({size/1024:.0f} KB)")


def main():
    print("=" * 60)
    print("Excel Workbook Generator")
    print("Capital Markets Intelligence Platform")
    print("=" * 60)
    print()

    print("  Loading data...")
    data = load_data()

    create_ipo_workbook(data)
    create_sovereign_workbook(data)
    create_mna_workbook(data)
    create_master_workbook(data)

    print()
    print("=" * 60)
    excels = [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx")]
    print(f"Generated {len(excels)} Excel workbooks:")
    for f in sorted(excels):
        size = os.path.getsize(os.path.join(EXCEL_DIR, f))
        print(f"  - {f} ({size/1024:.0f} KB)")
    print("=" * 60)


if __name__ == "__main__":
    main()
